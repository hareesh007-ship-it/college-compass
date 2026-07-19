#!/usr/bin/env python3
"""
Build US College Selection spreadsheet (29-column Investment Insights layout).

Outputs (per student profile):
  output/{FirstName} - US College Selection.xlsx
  output/{FirstName} - College Prep Gap Analysis.html

HS historical outcomes (Maia/Naviance) deferred — see docs/ENHANCEMENT.md.
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Any, Dict, List, Optional

from _paths import OUTPUT
from acceptance_data import display_pair
from college_finder import effective_act
from college_research import (
    accept_rate_program_column_label,
    program_rank_for_major,
    sort_rank_key,
    us_news_column_label,
)
from load_profile import profile_first_name
from output_paths import selection_output_paths, tuition_column_label
from fit_categories import fit_tier_excel_rows


def sheet_columns(profile: Dict[str, Any]) -> List[str]:
    """Column order: decision fields grouped by tuition and deadlines."""
    major = profile.get("intended_major", "")
    rank_col = us_news_column_label(major)
    tuition_col = tuition_column_label(profile)
    program_rate_col = accept_rate_program_column_label(major)
    applying_fall = profile.get("application_cycle", {}).get("applying_fall", 2026)
    apply_col = f"Apply Fall {applying_fall}"
    return [
        "University Name",
        "Public/Private",
        rank_col,
        tuition_col,
        "Avg Net Price (Scorecard)",
        "Fit Category",
        "Data Quality",
        "Test Optional",
        "Program Admit Type",
        "Program Requirements",
        "Student Meets Program Req?",
        "Program Admit Notes",
        "GPA Mid-50% Range",
        "GPA vs Mid-50%",
        "SAT Mid-50% Range",
        "SAT vs Mid-50% (effective)",
        "ACT Mid-50% Range",
        "ACT vs Mid-50% (effective)",
        "Early Action Deadline",
        "Early Decision Deadline",
        "Regular Decision Deadline",
        "ED Binding",
        "Within Budget",
        apply_col,
        "Accept Rate (University General)",
        program_rate_col,
        "Business Program Notes",
        "Accept Rate Source",
        "US News Rankings Source",
    ]

# Video CS sheet used general/OOS columns — replaced by business rates in data/acceptance_rates.json


def data_quality_label(entry: Dict[str, Any]) -> str:
    """High / Medium / Low based on how complete the admit stats are."""
    source = (entry.get("admit_stats_source") or "").lower()
    gpa_has = entry.get("gpa_range_display") not in (None, "", "—")
    sat_has = entry.get("sat_range_display") not in (None, "", "—")
    act_has = entry.get("act_range_display") not in (None, "", "—")
    bands = sum([gpa_has, sat_has, act_has])
    if bands >= 2 and "point estimate" not in source:
        return "High"
    if bands >= 1 or ("published" in source and "partial" in source):
        return "Medium"
    return "Low"


def fmt_pct(rate: Optional[float]) -> str:
    if rate is None:
        return ""
    return f"{rate * 100:.0f}%"


def fmt_money(amount: int) -> str:
    return f"${amount:,}"


def fmt_delta(value: float, suffix: str = "") -> str:
    sign = "+" if value >= 0 else ""
    return f"{sign}{value:.2f}{suffix}"


def apply_recommendation(entry: Dict[str, Any]) -> str:
    if entry.get("excluded"):
        return "No — excluded"
    cat = entry.get("category", "")
    if entry.get("priority_interest"):
        return f"Yes — {cat} (top interest)"
    if cat == "Safety":
        return "Yes — Safety"
    if cat == "Target":
        return "Yes — Target"
    if cat == "Reach":
        return "Consider — Reach"
    return "Review"


def build_row(
    entry: Dict[str, Any],
    profile: Dict[str, Any],
    eff_sat: int,
    eff_act: int,
    rank_col: str,
    tuition_col: str,
    program_rate_col: str,
    apply_col: str = "Apply Fall 2026",
) -> Dict[str, str]:
    accept_general, accept_business, accept_source = display_pair(entry["name"])
    if not accept_general and not accept_business:
        accept_general = fmt_pct(entry.get("acceptance_rate_used"))
        accept_source = "Fallback: college_compass estimate"

    dl = entry.get("deadlines", {})
    ed_binding = "Y" if dl.get("early_decision") and entry.get("early_decision_available") else "N"

    notes = ""
    if entry.get("exclusion_reasons"):
        notes = "; ".join(entry["exclusion_reasons"])
    elif entry.get("reasons"):
        for r in entry["reasons"]:
            if "business" in r.lower() or "entrepreneur" in r.lower() or "Carlson" in r or "Kelley" in r or "Tippie" in r or "Daniels" in r or "Gies" in r:
                notes = r
                break
        if not notes and len(entry["reasons"]) > 4:
            notes = entry["reasons"][4] if len(entry["reasons"]) > 4 else entry["reasons"][-1]

    rank_display, rank_note = program_rank_for_major(entry, profile.get("intended_major", ""))
    if rank_note:
        rank_display = f"{rank_display} ({rank_note})"

    row: Dict[str, str] = {
        "University Name": entry["name"],
        "Public/Private": entry["public_private"],
        rank_col: rank_display,
        tuition_col: fmt_money(entry["tuition_estimate"]),
        "Avg Net Price (Scorecard)": fmt_money(entry["avg_net_price"]) if entry.get("avg_net_price") else "—",
        "Fit Category": entry.get("category") or ("Excluded" if entry.get("excluded") else ""),
        "Data Quality": data_quality_label(entry),
        "Test Optional": "Yes" if entry.get("test_optional") else "No",
        "Program Admit Type": entry.get("program_admit_type", "—"),
        "Program Requirements": entry.get("program_requirements", "—"),
        "Student Meets Program Req?": entry.get("student_meets_program_req", "—"),
        "Program Admit Notes": entry.get("program_admit_notes", ""),
        "GPA Mid-50% Range": entry.get("gpa_range_display") or "—",
        "GPA vs Mid-50%": entry.get("gpa_position", "—"),
        "SAT Mid-50% Range": entry.get("sat_range_display") or "—",
        "SAT vs Mid-50% (effective)": entry.get("sat_position", "—"),
        "ACT Mid-50% Range": entry.get("act_range_display") or "—",
        "ACT vs Mid-50% (effective)": entry.get("act_position", "—"),
        "Early Action Deadline": dl.get("early_action") or "",
        "Early Decision Deadline": dl.get("early_decision") or "",
        "Regular Decision Deadline": dl.get("regular") or "",
        "ED Binding": ed_binding,
        "Within Budget": "Y" if entry.get("within_budget") else "N",
        apply_col: apply_recommendation(entry),
        "Accept Rate (University General)": accept_general,
        program_rate_col: accept_business,
        "Business Program Notes": notes,
        "Accept Rate Source": accept_source,
        "US News Rankings Source": entry.get("us_news_rankings_source", ""),
    }
    return row


def profile_header_rows(profile: Dict[str, Any], eff_sat: int, eff_act: int, columns: List[str]) -> List[List[str]]:
    pad = [""] * (len(columns) - 1)
    cycle = profile.get("application_cycle", {})
    major = profile.get("intended_major", "")
    rank_col = us_news_column_label(major)
    student_name = profile.get("name") or "Student"
    header: List[List[str]] = [
        [f"US College Selection — {student_name}"] + pad,
        [""] + pad,
        [
            "Intended Major",
            major,
            "",
            f"US News program rank: {rank_col}; Program Admit Type handles criteria schools (e.g. Iowa Tippie)",
        ] + pad[3:],
        ["High School", profile.get("high_school", ""), profile.get("high_school_ceeb", "")] + pad[3:],
        ["GPA (unweighted)", str(profile["gpa_unweighted"]), f"Weighted: {profile.get('gpa_weighted', '')}"] + pad[3:],
        ["SAT", str(profile.get("sat") or "")] + pad[2:],
    ]
    act = profile.get("act")
    if isinstance(act, dict):
        header.append([
            "ACT",
            str(act.get("composite")),
            f"E{act.get('english')}/M{act.get('math')}/R{act.get('reading')}/S{act.get('science')}",
            f"Effective SAT: {eff_sat} | Effective ACT: {eff_act}",
        ] + pad[4:])
    else:
        header.append(["ACT", str(act or ""), f"Effective SAT: {eff_sat} | Effective ACT: {eff_act}"] + pad[3:])
    header.extend([
        ["State of Residence", profile["state_of_residence"]] + pad[2:],
        ["Grade / Cycle", f"Grade {profile['grade']}", f"Apply Fall {cycle.get('applying_fall', 2026)} → Start {cycle.get('college_start', '2027-08')}"] + pad[3:],
        ["Max Tuition Budget", fmt_money(profile["budget_max_tuition_per_year"])] + pad[2:],
    ])
    header.extend(fit_tier_excel_rows(len(columns)))
    return header


def sort_entries(report: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Recommended schools first, then excluded; user-preferred schools first within matches."""
    order = {"Safety": 0, "Target": 1, "Reach": 2, "Excluded": 3}

    def key(e: Dict[str, Any]) -> tuple:
        if e.get("excluded"):
            cat = 3
        else:
            cat = order.get(e.get("category", ""), 9)
        return (cat, not e.get("priority_interest", False), sort_rank_key(e, report["student"].get("intended_major", "")))

    combined = list(report["matches"]) + list(report["excluded"])
    combined.sort(key=key)
    return combined


def build_checklist_rows(report: Dict[str, Any], applying_fall: int) -> List[List[str]]:
    headers = [
        "University",
        "Fit",
        "Apply?",
        "Platform",
        "Early Action",
        "Early Decision",
        "Regular Decision",
        "Fall 2026 Task",
        "Done",
    ]
    rows: List[List[str]] = [headers]

    global_tasks = [
        ["— ALL SCHOOLS —", "", "", "Common App", "", "", "", f"Create Common App account (Aug {applying_fall})", ""],
        ["— ALL SCHOOLS —", "", "", "Common App", "", "", "", f"Draft personal statement (Aug–Sep {applying_fall})", ""],
        ["— ALL SCHOOLS —", "", "", "Common App", "", "", "", f"Request counselor recommendations (Sep {applying_fall})", ""],
        ["— ALL SCHOOLS —", "", "", "Testing", "", "", "", f"Optional SAT retake (Sep–Oct {applying_fall})", ""],
    ]
    rows.extend(global_tasks)

    for entry in report["matches"]:
        if entry.get("excluded"):
            continue
        dl = entry.get("deadlines", {})
        ea = dl.get("early_action")
        ed = dl.get("early_decision")
        reg = dl.get("regular")

        def date_str(label: Optional[str]) -> str:
            if not label:
                return ""
            if "Nov 1" in label:
                return f"Nov 1, {applying_fall}"
            if "Nov 15" in label:
                return f"Nov 15, {applying_fall}"
            if "Jan 1" in label:
                return f"Jan 1, {applying_fall + 1}"
            if "Jan 5" in label:
                return f"Jan 5, {applying_fall + 1}"
            if "Jan 15" in label:
                return f"Jan 15, {applying_fall + 1}"
            if "Feb 1" in label:
                return f"Feb 1, {applying_fall + 1}"
            if "May 1" in label:
                return f"May 1, {applying_fall + 1}"
            return label

        task = "Submit application"
        if ea:
            task = f"Submit EA by {date_str(ea)}"
        elif ed:
            task = f"Submit ED by {date_str(ed)} (binding — confirm first choice)"
        else:
            task = f"Submit by {date_str(reg)}"

        rows.append([
            entry["name"],
            entry.get("category", ""),
            apply_recommendation(entry),
            "Common App",
            date_str(ea),
            date_str(ed),
            date_str(reg),
            task,
            "",
        ])

    return rows


from build_gap_analysis import write_gap_analysis


def _sheet_tab_name(major: str) -> str:
    """Build a tab name like 'College Selection - Business', capped at 31 chars."""
    prefix = "College Selection - "
    available = 31 - len(prefix)
    label = (major.split()[0] if major.split() else major)[:available]
    return prefix + label


def write_xlsx(
    path: str,
    tab_title: str,
    header_block: List[List[str]],
    data_rows: List[Dict[str, str]],
    columns: List[str],
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("openpyxl not installed — skipping .xlsx (CSV still written)")
        return

    wb = Workbook()

    header_font = Font(bold=True)
    section_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws = wb.active
    ws.title = tab_title[:31]

    for row in header_block:
        ws.append(row)
    ws.append(columns)

    header_row_num = len(header_block) + 1
    for cell in ws[header_row_num]:
        cell.font = header_font
        cell.fill = section_fill

    for row in data_rows:
        ws.append([row.get(col, "") for col in columns])

    for cell in ws[4]:
        cell.font = Font(bold=True)

    ws.freeze_panes = f"A{len(header_block) + 2}"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["Y"].width = 48
    ws.column_dimensions["Z"].width = 36

    wb.save(path)


def write_selection_outputs(
    profile: Dict[str, Any],
    report: Dict[str, Any],
    profile_source: str,
) -> int:
    """Build XLSX + gap analysis HTML from an existing profile + match report."""
    eff_sat = report["matching_notes"]["effective_sat_used"]
    eff_act = report["matching_notes"]["effective_act_used"]
    paths = selection_output_paths(profile)

    columns = sheet_columns(profile)
    rank_col = us_news_column_label(profile.get("intended_major", ""))
    tuition_col = tuition_column_label(profile)
    program_rate_col = accept_rate_program_column_label(profile.get("intended_major", ""))
    applying_fall = profile.get("application_cycle", {}).get("applying_fall", 2026)
    apply_col = f"Apply Fall {applying_fall}"
    header = profile_header_rows(profile, eff_sat, eff_act, columns)
    entries = sort_entries(report)
    data_rows = [
        build_row(e, profile, eff_sat, eff_act, rank_col, tuition_col, program_rate_col, apply_col) for e in entries
    ]

    gap_html = write_gap_analysis(profile, report)
    tab_title = _sheet_tab_name(profile.get("intended_major", ""))

    os.makedirs(OUTPUT, exist_ok=True)
    write_xlsx(str(paths["xlsx"]), tab_title, header, data_rows, columns)

    print(f"Wrote {paths['xlsx']} (if openpyxl available)")
    print(
        f"Rows: {len(data_rows)} colleges "
        f"({len(report['matches'])} recommended, {len(report['excluded'])} reference/excluded)"
    )
    try:
        from run_log import log_sheet_build  # noqa: WPS433

        log_sheet_build(
            profile_path=profile_source,
            rows=len(data_rows),
            outputs=[paths["xlsx"].name, gap_html.name],
        )
    except Exception:
        pass
    return len(data_rows)


def main() -> int:
    from pipeline import run_pipeline

    return run_pipeline(write_matches=True, write_sheet=True)


if __name__ == "__main__":
    raise SystemExit(main())
