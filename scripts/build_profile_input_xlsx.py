#!/usr/bin/env python3
"""Build input/student profile input.xlsx — single-tab student profile form."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from _paths import INPUT
from profile_fields import PROFILE_SECTIONS, PROFILE_XLSX

AUTHOR = "College Finder"

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
REQUIRED_FILL = PatternFill("solid", fgColor="FFF2CC")
SECTION_FONT = Font(bold=True, size=11)
REQUIRED_FONT = Font(bold=True)
INSTRUCTION_FONT = Font(italic=True, size=10)


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Profile"

    ws["A1"] = "College Finder — Student Profile Input"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")

    ws["A2"] = (
        "Fill REQUIRED fields (yellow). Save this file in input/. "
        "Add transcript.pdf, resume.pdf, or test reports in the same folder when helpful."
    )
    ws["A2"].font = INSTRUCTION_FONT
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:B2")
    ws.row_dimensions[2].height = 36

    ws["A3"] = "Field"
    ws["B3"] = "Your answer"
    for col in ("A", "B"):
        cell = ws[f"{col}3"]
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A4"

    row = 4
    for section_title, fields in PROFILE_SECTIONS:
        ws.cell(row=row, column=1, value=section_title)
        ws.cell(row=row, column=1).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="E7E6E6")
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="E7E6E6")
        row += 1

        for spec in fields:
            label_cell = ws.cell(
                row=row, column=1, value=f"{spec.label} *" if spec.required else spec.label
            )
            value_cell = ws.cell(row=row, column=2, value=spec.example or "")
            value_cell.alignment = Alignment(wrap_text=True, vertical="top")

            if spec.required:
                label_cell.font = REQUIRED_FONT
                label_cell.fill = REQUIRED_FILL
                value_cell.fill = REQUIRED_FILL

            if spec.comment:
                label_cell.comment = Comment(spec.comment, AUTHOR, width=320, height=120)

            row += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 64

    return wb


def main() -> int:
    INPUT.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(PROFILE_XLSX)
    print(f"Wrote {PROFILE_XLSX}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
