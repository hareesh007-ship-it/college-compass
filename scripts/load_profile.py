"""Load student profile from Excel, then enrich with data from PDFs in input/."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from openpyxl import load_workbook

from profile_fields import (
    PROFILE_XLSX,
    ProfileFieldSpec,
    field_by_label,
)

MATCHER_CRITICAL_PATHS = ("gpa_unweighted",)


class ProfileLoadError(ValueError):
    pass


def _normalize_label(raw: Any) -> str:
    text = str(raw or "").strip()
    if text.endswith("*"):
        text = text[:-1].strip()
    return text


def _parse_bool(raw: str) -> bool:
    value = raw.strip().lower()
    if value in {"yes", "y", "true", "1"}:
        return True
    if value in {"no", "n", "false", "0"}:
        return False
    raise ProfileLoadError(f"Expected Yes or No, got {raw!r}")


def _split_list(raw: str, sep: str) -> List[str]:
    if sep == ",":
        parts = raw.split(",")
    else:
        parts = raw.split(";")
    return [p.strip() for p in parts if p.strip()]


def _coerce_value(spec: ProfileFieldSpec, raw: Any) -> Any:
    if raw is None:
        return None
    if isinstance(raw, bool):
        text = "Yes" if raw else "No"
    elif isinstance(raw, (int, float)):
        text = str(raw)
    else:
        text = str(raw).strip()
    if not text:
        return None

    kind = spec.kind
    if kind == "str":
        return text
    if kind == "int":
        try:
            return int(float(text.replace(",", "")))
        except ValueError as exc:
            raise ProfileLoadError(f"{spec.label}: expected a whole number, got {text!r}") from exc
    if kind == "float":
        try:
            return float(text)
        except ValueError as exc:
            raise ProfileLoadError(f"{spec.label}: expected a number, got {text!r}") from exc
    if kind == "bool":
        return _parse_bool(text)
    if kind == "comma_list":
        return _split_list(text, ",")
    if kind == "semicolon_list":
        return _split_list(text, ";")
    return text


def _set_path(profile: Dict[str, Any], path: str, value: Any) -> None:
    parts = path.split(".")
    node = profile
    for part in parts[:-1]:
        child = node.get(part)
        if not isinstance(child, dict):
            child = {}
            node[part] = child
        node = child
    node[parts[-1]] = value


def _get_path(profile: Dict[str, Any], path: str) -> Any:
    node: Any = profile
    for part in path.split("."):
        if not isinstance(node, dict) or part not in node:
            return None
        node = node[part]
    return node


def _finalize_act(profile: Dict[str, Any]) -> None:
    act = profile.get("act")
    if not isinstance(act, dict):
        profile.pop("act", None)
        return
    if act.get("composite") is None:
        profile.pop("act", None)
        return
    cleaned = {k: v for k, v in act.items() if v is not None}
    profile["act"] = cleaned


def _finalize_campus_size(profile: Dict[str, Any]) -> None:
    prefs = profile.get("preferences")
    if not isinstance(prefs, dict):
        return
    size = prefs.get("campus_size")
    if size == "" or size is None:
        prefs["campus_size"] = None


def _validate_profile(profile: Dict[str, Any], specs: Dict[str, ProfileFieldSpec]) -> None:
    missing: List[str] = []
    for spec in specs.values():
        if not spec.required:
            continue
        if _get_path(profile, spec.path) in (None, "", []):
            missing.append(spec.label)
    if missing:
        raise ProfileLoadError(
            "Missing required profile fields in Excel: " + ", ".join(missing)
        )

    for path in MATCHER_CRITICAL_PATHS:
        if _get_path(profile, path) is None:
            raise ProfileLoadError(
                f"Matcher needs {path.replace('.', ' → ')} — fill Unweighted GPA in Excel "
                "(or add transcript.pdf and enter GPA when known)."
            )

    prefs = profile.setdefault("preferences", {})
    prefs.setdefault("public_ok", True)
    prefs.setdefault("private_ok", True)
    prefs.setdefault("regions", [])


def load_profile_from_xlsx(path: Union[str, Path]) -> Dict[str, Any]:
    path = Path(path)
    if not path.is_file():
        raise ProfileLoadError(f"Profile Excel not found: {path}")

    specs = field_by_label()
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Student Profile" not in wb.sheetnames:
            raise ProfileLoadError(f"{path}: missing sheet 'Student Profile'")
        ws = wb["Student Profile"]
        profile: Dict[str, Any] = {}

        for row in ws.iter_rows(min_row=4, max_col=2, values_only=True):
            label_raw, value = row[0], row[1] if len(row) > 1 else None
            if not label_raw:
                continue
            label = _normalize_label(label_raw)
            spec = specs.get(label)
            if spec is None:
                continue
            parsed = _coerce_value(spec, value)
            if parsed is not None:
                _set_path(profile, spec.path, parsed)
    finally:
        wb.close()

    _finalize_act(profile)
    _finalize_campus_size(profile)
    _validate_profile(profile, specs)
    profile["_source"] = str(path)
    return profile


def load_profile(path: Optional[Union[str, Path]] = None) -> Dict[str, Any]:
    """Load profile from Excel, then fill blank fields from PDFs in input/."""
    profile_path = Path(path) if path is not None else PROFILE_XLSX
    suffix = profile_path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise ProfileLoadError(
            f"Profile must be an Excel file (.xlsx): {profile_path}. "
            f"Use {PROFILE_XLSX.name}."
        )
    profile = load_profile_from_xlsx(profile_path)

    # Enrich from PDFs (transcript.pdf, resume.pdf, etc.) if present
    input_dir = profile_path.parent
    try:
        from pro_config import load_pro_config
        backend = load_pro_config().get("research_backend", "cursor")
    except Exception:
        backend = "cursor"

    if backend != "cursor":
        try:
            from extract_student_docs import extract_and_merge
            n = extract_and_merge(profile, input_dir, backend)
            if n:
                print(f"  [student docs] Filled {n} profile fields from documents in {input_dir}")
        except Exception as exc:
            print(f"  [student docs] WARNING: PDF extraction failed: {exc}")

    return profile


def profile_first_name(profile: Dict[str, Any]) -> str:
    name = str(profile.get("name", "Student")).strip()
    return name.split()[0] if name else "Student"
